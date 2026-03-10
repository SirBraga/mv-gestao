from collections import defaultdict
import csv
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

SRC = Path('/Users/sirbraga/Downloads/clientes.xlsx')
OUT_DIR = Path('/Users/sirbraga/Downloads/clientes_seed_csv')
OUT_DIR.mkdir(parents=True, exist_ok=True)

BUSINESS_OPTIONS = [
    'Comércio',
    'Serviços',
    'Indústria',
    'Construção Civil',
    'Tecnologia da Informação',
    'Saúde',
    'Educação',
    'Transporte',
    'Alimentação',
    'Consultoria',
    'Agricultura',
    'Pecuária',
    'Mineração',
    'Energia',
    'Saneamento Básico',
    'Telecomunicações',
    'Mídia e Comunicação',
]

BUSINESS_KEYWORDS = {
    'Alimentação': ['restaurante', 'lanchonete', 'pizzaria', 'food', 'bar', 'acai', 'sorveteria', 'padaria', 'confeitaria', 'hamburgueria', 'alimentacao'],
    'Comércio': ['comercio', 'loja', 'mercado', 'supermercado', 'varejo', 'atacado', 'farmacia', 'papelaria', 'calcados', 'roupas', 'auto pecas'],
    'Serviços': ['servico', 'serviços', 'oficina', 'salao', 'salão', 'manutencao', 'manutenção', 'prestacao'],
    'Tecnologia da Informação': ['software', 'informatica', 'informática', 'tecnologia', 'ti', 'sistema'],
    'Saúde': ['clinica', 'clínica', 'hospital', 'saude', 'saúde', 'odonto', 'laboratorio'],
    'Educação': ['escola', 'colegio', 'colégio', 'educacao', 'educação', 'curso'],
    'Transporte': ['transporte', 'logistica', 'logística', 'frete'],
    'Consultoria': ['consultoria', 'contabilidade', 'assessoria'],
    'Construção Civil': ['construcao', 'construção', 'material de construcao', 'engenharia'],
    'Indústria': ['industria', 'indústria', 'fabrica', 'fábrica'],
    'Agricultura': ['agricola', 'agrícola', 'fazenda', 'agro'],
    'Pecuária': ['pecuaria', 'pecuária', 'gado'],
    'Mineração': ['mineracao', 'mineração', 'mineradora'],
    'Energia': ['energia', 'solar', 'eletrica', 'elétrica'],
    'Saneamento Básico': ['saneamento', 'agua', 'água', 'esgoto'],
    'Telecomunicações': ['telecom', 'internet', 'provedor'],
    'Mídia e Comunicação': ['midia', 'mídia', 'marketing', 'publicidade', 'comunicacao', 'comunicação'],
}

PRODUCT_RENAMES = {
    'HOTLINE': 'HOST',
    'HOST': 'HOST',
    'CPLUG': 'Connect Plug',
    'CLIPP PRO': 'Clipp PRO',
    'GRFOOD': 'GR Food',
    'COMANDA10': 'Comanda 10',
    'HIPER': 'Híper',
    'CLIPP 360': 'Clipp 360',
    'CLIPP FACIL': 'Clipp Fácil',
    'BACKUP': 'Backup',
    'MANUNTENÇÃO': 'Manutenção',
    'MANUTENÇÃO': 'Manutenção',
    'MANUTENCAO': 'Manutenção',
}

ZWEB_PLUGIN_MAP = {
    'SINTEGRA': 'Plugin Sintegra',
    'VENDA GERENCIAL': 'Plugin Venda Gerencial',
    'VENDAGERENCIAL': 'Plugin Venda Gerencial',
    'INTEGRA': 'Plugin Integra',
    'IFOOD': 'Plugin iFood',
    'DELIVERY': 'Plugin Delivery',
    'PIX': 'Plugin Pix',
    'BALANCA': 'Plugin Balança',
    'BALANÇA': 'Plugin Balança',
}

SERIAL_PRODUCTS = {'ZWEB Standard', 'ZWEB Premium', 'ZWEB Essential'}
SERIAL_EXPIRES = '2026-12-31'
MV_BASE = 'Sistema MV'


def normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = text.replace('\n', ' ').replace('\r', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text


def strip_emoji(text):
    return ''.join(ch for ch in text if unicodedata.category(ch) != 'So')


def titleish(text):
    text = strip_emoji(normalize_text(text))
    if not text:
        return ''
    words = []
    keep_lower = {'da', 'de', 'do', 'das', 'dos', 'e'}
    for i, part in enumerate(text.split(' ')):
        lower = part.lower()
        if i > 0 and lower in keep_lower:
            words.append(lower)
        else:
            words.append(lower[:1].upper() + lower[1:])
    return ' '.join(words)


def digits(value):
    return re.sub(r'\D+', '', normalize_text(value))


def clean_phone(value):
    raw = digits(value)
    if raw in {'', '0'}:
        return ''
    return raw


def clean_obs(value):
    text = strip_emoji(normalize_text(value))
    text = re.sub(r'\s*[,;|/]+\s*', ' | ', text)
    text = re.sub(r'(\s*\|\s*){2,}', ' | ', text)
    return text.strip(' |')


def plain(value):
    return unicodedata.normalize('NFD', normalize_text(value)).encode('ascii', 'ignore').decode('ascii').lower()


def normalize_business_sector(raw):
    text = normalize_text(raw)
    if not text:
        return 'Serviços'
    base = plain(text)
    for option, keywords in BUSINESS_KEYWORDS.items():
        if any(keyword in base for keyword in keywords):
            return option
    for option in BUSINESS_OPTIONS:
        if plain(option) in base:
            return option
    return 'Serviços'


def normalize_product_name(name):
    raw = normalize_text(name)
    upper = raw.upper()
    if upper in PRODUCT_RENAMES:
        return PRODUCT_RENAMES[upper]
    if upper.startswith('ZWEB STANDARD'):
        return 'ZWEB Standard'
    if upper.startswith('ZWEB PREMIUM'):
        return 'ZWEB Premium'
    if upper.startswith('ZWEB ESSENTIAL'):
        return 'ZWEB Essential'
    if upper.startswith('SISTEMA MV'):
        return MV_BASE
    return titleish(raw)


def split_multi(value):
    text = normalize_text(value)
    if not text:
        return []
    temp = re.sub(r'\s*/\s*', ' | ', text)
    temp = re.sub(r'\s*;\s*', ' | ', temp)
    return [part.strip() for part in temp.split('|') if part.strip()]


def extract_products_and_plugins(raw_value):
    bases = []
    plugins_by_base = defaultdict(list)
    installations = {}
    for item in split_multi(raw_value):
        upper = item.upper()
        if upper.startswith('ZWEB PLUG'):
            suffix = upper.replace('ZWEB PLUG', '').strip()
            plugin_name = ZWEB_PLUGIN_MAP.get(suffix, titleish('Plugin ' + suffix.lower()))
            base = 'ZWEB Standard'
            if base not in bases:
                bases.append(base)
            if plugin_name not in plugins_by_base[base]:
                plugins_by_base[base].append(plugin_name)
            continue
        normalized = normalize_product_name(item)
        if normalized == MV_BASE:
            if 'LOCAL' in upper:
                installations[normalized] = 'LOCAL'
            elif 'SERVIDOR' in upper:
                installations[normalized] = 'SERVIDOR'
            else:
                installations[normalized] = 'ONLINE'
        if normalized not in bases:
            bases.append(normalized)
    return bases, plugins_by_base, installations


def detect_sheet(workbook):
    if 'dados_limpos' in workbook.sheetnames:
        return workbook['dados_limpos']
    return workbook[workbook.sheetnames[0]]


def resolve_columns(headers):
    aliases = {
        'nome_fantasia': ['Nome Fantasia'],
        'razao_social': ['Razão Social', 'Razao Social'],
        'cpf_cnpj': ['CPF/CNPJ'],
        'ie': ['Inscrição Estadual', 'Inscricao Estadual'],
        'cnae': ['CNAE', 'CNAE Principal'],
        'ramo': ['Ramo de Atividade', 'Ramo Atividade', 'Atividade Principal'],
        'produto': ['Produto', 'Produtos', 'Sistema', 'Serviço', 'Servico', 'Serviços Contratados', 'Servicos Contratados'],
        'serial': ['Serial', 'Seriais'],
        'valor': ['Valor', 'Valor Mensal', 'Mensalidade'],
        'observacoes': ['Observações', 'Observacoes'],
        'contabilidade': ['Contabilidade', 'Escritório Contábil', 'Escritorio Contabil', 'Contador'],
        'cidade': ['Cidade'],
        'endereco': ['Endereço', 'Endereco'],
        'numero': ['Número', 'Numero'],
        'bairro': ['Bairro'],
        'cep': ['CEP'],
        'complemento': ['Complemento'],
        'telefone': ['Telefone', 'Telefone Principal', 'Contato - Número'],
        'email': ['Email', 'E-mail', 'Contato - E-mail'],
        'owner_name': ['Nome Proprietário', 'Nome Proprietario', 'Proprietário', 'Proprietario'],
        'owner_phone': ['Telefone Proprietário', 'Telefone Proprietario'],
        'owner_email': ['Email Proprietário', 'Email Proprietario'],
        'owner_cpf': ['CPF Proprietário', 'CPF Proprietario'],
        'extra_name': ['Contato Extra', 'Nome Contato Extra', 'Contato - Nome'],
        'extra_phone': ['Telefone Contato Extra'],
        'extra_email': ['Email Contato Extra'],
        'extra_role': ['Cargo Contato Extra', 'Função Contato Extra', 'Funcao Contato Extra'],
    }
    index = {normalize_text(header): i for i, header in enumerate(headers)}
    resolved = {}
    for key, names in aliases.items():
        for name in names:
            if name in index:
                resolved[key] = index[name]
                break
    return resolved


def get_cell(row, columns, key):
    idx = columns.get(key)
    if idx is None or idx >= len(row):
        return ''
    return normalize_text(row[idx])


def write_csv(path, rows, headers):
    with path.open('w', newline='', encoding='utf-8-sig') as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, '') for header in headers})


def main():
    workbook = load_workbook(SRC, read_only=True, data_only=True)
    worksheet = detect_sheet(workbook)
    iterator = worksheet.iter_rows(values_only=True)
    headers = [normalize_text(value) for value in next(iterator)]
    columns = resolve_columns(headers)

    clients = []
    client_products = []
    client_product_plugins = []
    client_serials = []

    for row_num, row in enumerate(iterator, start=2):
        razao = titleish(get_cell(row, columns, 'razao_social'))
        fantasia = titleish(get_cell(row, columns, 'nome_fantasia'))
        document = digits(get_cell(row, columns, 'cpf_cnpj'))
        if not (razao or fantasia or document):
            continue

        is_pj = len(document) > 11 or not document
        client_name = fantasia or razao or titleish(get_cell(row, columns, 'nome_fantasia') or get_cell(row, columns, 'razao_social'))
        client_key = document or f'ROW-{row_num}'
        notes = clean_obs(get_cell(row, columns, 'observacoes'))
        amount = normalize_text(get_cell(row, columns, 'valor')).replace('R$', '').strip()
        bases, plugins_by_base, installations = extract_products_and_plugins(get_cell(row, columns, 'produto'))
        serials = split_multi(get_cell(row, columns, 'serial'))
        contact_name = titleish(get_cell(row, columns, 'extra_name'))
        contact_phone = clean_phone(get_cell(row, columns, 'extra_phone') or get_cell(row, columns, 'telefone'))
        contact_email = get_cell(row, columns, 'extra_email').lower() or get_cell(row, columns, 'email').lower()
        owner_name = titleish(get_cell(row, columns, 'owner_name')) or contact_name
        owner_phone = clean_phone(get_cell(row, columns, 'owner_phone')) or contact_phone
        owner_email = get_cell(row, columns, 'owner_email').lower() or contact_email

        clients.append({
            'client_key': client_key,
            'type': 'PJ' if is_pj else 'PF',
            'name': client_name if is_pj else (client_name or razao),
            'razaoSocial': razao if is_pj else '',
            'nomeFantasia': fantasia if is_pj else '',
            'cnpj': document if is_pj else '',
            'cpf': '' if is_pj else document,
            'ie': digits(get_cell(row, columns, 'ie')),
            'cnae': digits(get_cell(row, columns, 'cnae')),
            'businessSector': normalize_business_sector(get_cell(row, columns, 'ramo') or client_name),
            'address': titleish(get_cell(row, columns, 'endereco')),
            'houseNumber': get_cell(row, columns, 'numero'),
            'neighborhood': titleish(get_cell(row, columns, 'bairro')),
            'city': titleish(get_cell(row, columns, 'cidade')),
            'zipCode': digits(get_cell(row, columns, 'cep')),
            'complement': titleish(get_cell(row, columns, 'complemento')),
            'ownerName': owner_name,
            'ownerPhone': owner_phone,
            'ownerEmail': owner_email,
            'ownerCpf': digits(get_cell(row, columns, 'owner_cpf')),
            'phone': clean_phone(get_cell(row, columns, 'telefone')),
            'email': get_cell(row, columns, 'email').lower(),
            'aditionalInfo': notes,
            'hasContract': 'true',
            'contractType': 'MENSAL',
            'supportReleased': 'false',
            'certificateType': '',
            'certificateExpiresDate': '',
            'contabilityRaw': titleish(get_cell(row, columns, 'contabilidade')),
            'extraContactName': contact_name,
            'extraContactPhone': contact_phone,
            'extraContactEmail': contact_email,
            'extraContactRole': titleish(get_cell(row, columns, 'extra_role')) or 'Colaborador',
        })

        for index_base, base in enumerate(bases, start=1):
            client_product_key = f'{client_key}::{base}'
            client_products.append({
                'client_key': client_key,
                'client_product_key': client_product_key,
                'productName': base,
                'installationType': installations.get(base, ''),
                'priceMonthly': amount,
                'priceQuarterly': '',
                'priceYearly': '',
                'notes': notes,
                'hasSerialControl': 'true' if base in SERIAL_PRODUCTS else 'false',
            })

            for plugin_name in plugins_by_base.get(base, []):
                client_product_plugins.append({
                    'client_product_key': client_product_key,
                    'productName': base,
                    'pluginName': plugin_name,
                    'priceMonthly': '',
                    'priceQuarterly': '',
                    'priceYearly': '',
                    'notes': '',
                })

            if base in SERIAL_PRODUCTS:
                serial_value = serials[index_base - 1] if index_base - 1 < len(serials) else ''
                client_serials.append({
                    'client_product_key': client_product_key,
                    'client_key': client_key,
                    'productName': base,
                    'serial': serial_value,
                    'expiresAt': SERIAL_EXPIRES,
                })

    client_headers = list(clients[0].keys()) if clients else ['client_key']
    product_headers = list(client_products[0].keys()) if client_products else ['client_product_key']
    plugin_headers = list(client_product_plugins[0].keys()) if client_product_plugins else ['client_product_key']
    serial_headers = list(client_serials[0].keys()) if client_serials else ['client_product_key']

    write_csv(OUT_DIR / 'clients.csv', clients, client_headers)
    write_csv(OUT_DIR / 'client_products.csv', client_products, product_headers)
    write_csv(OUT_DIR / 'client_product_plugins.csv', client_product_plugins, plugin_headers)
    write_csv(OUT_DIR / 'client_serials.csv', client_serials, serial_headers)

    print(f'CSV gerados em: {OUT_DIR}')
    print(f'Clientes: {len(clients)}')
    print(f'Produtos contratados: {len(client_products)}')
    print(f'Plugins contratados: {len(client_product_plugins)}')
    print(f'Seriais: {len(client_serials)}')


if __name__ == '__main__':
    main()
